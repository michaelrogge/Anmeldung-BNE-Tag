import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import qrcode
from functools import wraps
from flask import session, redirect, url_for

def export_to_csv(registrations):
    """Exportiert Anmeldungen zu CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Vorname', 'Nachname', 'Email', 'Telefon', 'Organisation',
        'Besondere Wünsche', 'Anmeldedatum', 'Workshops'
    ])
    
    # Daten
    for reg in registrations:
        participant = reg.participant
        workshops = ', '.join([wr.workshop.name for wr in reg.workshop_selections])
        
        writer.writerow([
            participant.first_name,
            participant.last_name,
            participant.email,
            participant.phone or '',
            participant.organization or '',
            participant.dietary_restrictions or '',
            reg.registration_date.strftime('%d.%m.%Y %H:%M'),
            workshops
        ])
    
    return output.getvalue()

def export_to_excel(registrations):
    """Exportiert Anmeldungen zu Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Anmeldungen"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Header
    headers = ['Vorname', 'Nachname', 'Email', 'Telefon', 'Organisation', 
               'Besondere Wünsche', 'Anmeldedatum', 'Workshops']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Daten
    for row, reg in enumerate(registrations, 2):
        participant = reg.participant
        workshops = ', '.join([wr.workshop.name for wr in reg.workshop_selections])
        
        ws.cell(row=row, column=1).value = participant.first_name
        ws.cell(row=row, column=2).value = participant.last_name
        ws.cell(row=row, column=3).value = participant.email
        ws.cell(row=row, column=4).value = participant.phone or ''
        ws.cell(row=row, column=5).value = participant.organization or ''
        ws.cell(row=row, column=6).value = participant.dietary_restrictions or ''
        ws.cell(row=row, column=7).value = reg.registration_date.strftime('%d.%m.%Y %H:%M')
        ws.cell(row=row, column=8).value = workshops
    
    # Spaltenbreite anpassen
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    
    # In Bytes speichern
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def generate_qr_code(url):
    """Generiert QR-Code aus URL"""
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Als Base64 für HTML
    import base64
    output = io.BytesIO()
    img.save(output, format="PNG")
    output.seek(0)
    img_base64 = base64.b64encode(output.getvalue()).decode()
    
    return f"data:image/png;base64,{img_base64}"

def admin_required(f):
    """Dekorator für Admin-Authentifizierung"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

def validate_email(email):
    """Validiert Email-Format"""
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def get_registration_summary(event):
    """Gibt Zusammenfassung der Anmeldungen"""
    from models import Registration
    
    registrations = Registration.query.filter_by(event_id=event.id).all()
    
    return {
        'total_participants': len(registrations),
        'confirmed': sum(1 for r in registrations if r.status == 'confirmed'),
        'registered': sum(1 for r in registrations if r.status == 'registered'),
        'cancelled': sum(1 for r in registrations if r.status == 'cancelled'),
    }
